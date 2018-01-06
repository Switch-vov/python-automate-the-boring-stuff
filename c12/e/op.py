import openpyxl

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))
print(workbook.get_sheet_names())
sheet = workbook.get_sheet_by_name('Sheet1')
print(type(sheet))
print(sheet.title)
active_sheet = workbook.get_active_sheet()
print(sheet.title)
A1 = active_sheet['A1']
print(A1)
print(A1.value)
print(type(A1.value))
print(A1.coordinate)
B1 = active_sheet.cell(row=1, column=2)
print(B1)
for i in range(1, 8):
    print(i, active_sheet.cell(row=i, column=2).value)
print(active_sheet.max_row)
print(active_sheet.max_column)
